import yaml
import collections
import openpyxl
import os
import datetime
import zk
import itertools
import calendar
import io
from win32com import client
from cryptography.fernet import Fernet
from .database import Database
from .biometric import Biometric
from PyPDF2 import PdfWriter, PdfReader
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

class DTR:
    def __init__(self, config:dict|str):
        '''
        Initialize DTR object

        `config` - if string is given tries to find the config file (should be a valid yml file)\n
        Config Format: \n
        config = { \n
            'database': { \n
                'host': '127.0.0.1', \n
                'database': 'dtr', \n
                'user': 'dtr', \n
                'password': 'dtr' \n
            }, \n
            'directory': { \n
                'export': 'exports' \n
            } \n
        } 

        `export_folder` - folder name for DTR exports
        '''
        self.__key = b'LeaO6Y59ngYo0eYA54b5VevDcQe0AOFJ7Whl1D7dt7c='
        cipher_suite = Fernet(self.__key)

        if(isinstance(config, dict)):
            config_data = config
        elif(isinstance(config, str)):
            if(config.split('.')[-1]) != 'yml':
                raise Exception('Config is not a valid yml file')
            with open(config, 'r') as file:
                config_data = yaml.safe_load(file)
        self.__dbhost = config_data['database']['host']
        self.__dbname = config_data['database']['database']
        self.__dbuser = config_data['database']['user']
        self.export_folder = config_data['directory']['export']
        self.__dbpass = cipher_suite.decrypt(config_data['database']['password']).decode()
        try:
            self.__database = Database(self.__dbhost, self.__dbname, self.__dbuser, self.__dbpass)
        except:
            raise Exception('Error connecting to database')
        self._directory = os.path.dirname(os.path.abspath(__file__))
        self.__dbverified = False
        self.__logged = False
        self.__id = None
        self.__username = ''
        self.__name = ''
        self.__level = None
        self.busy = False

    def update_config_file(self, config_file:str, dbhost:str, dbname:str, dbuser:str, dbpass:str, export_directory:str):
        '''
        Update a given config file (should be in a yml format)

        Returns `True` if successfully updated. Otherwise `False` \n
        Raises an Exception if config_file is not a valid yml file
        '''
        try:
            with open(config_file, 'r') as file:
                yaml.safe_load(file)
        except:
            raise Exception('Not a valid yml file')
        cipher_suite = Fernet(self.__key)
        config_data = {}
        config_data['database'] = {}
        config_data['directory'] = {}
        config_data['database']['host'] = dbhost
        config_data['database']['database'] = dbname
        config_data['database']['user'] = dbuser
        config_data['database']['password'] = cipher_suite.encrypt(dbpass.encode())
        config_data['directory']['export'] = export_directory
        with open(config_file, 'w') as file:
            yaml.dump(config_data, file)
        return True

    def __get_dtr(self, employee_id:int|str, month:int, year:int, cut:str = 'full'):
        '''
        Get DTR of employee on a certain month and year

        Returns a list of DTRs \n
        Raises and Exception if `employee_id` not found \n
        `cut` - cuts retrieved dtr by half of month \n
        Possible values:
         - `full` - whole month
         - `first` - first half of month
         - `last` - last half of month
        '''
        if(not self.__database.employee_exists(employee_id)):
            raise Exception('Employee not found')
        dtrs = []
        cut = cut.lower()
        if cut == 'full':
            query = 'SELECT * FROM dtrs WHERE employee_id = %s AND YEAR(date) = %s AND MONTH(date) = %s'
            values = (employee_id, year, month)
        elif cut == 'first':
            query = 'SELECT * FROM dtrs WHERE employee_id = %s AND YEAR(date) = %s AND MONTH(date) = %s AND DAY(date) < 16'
            values = (employee_id, year, month)
        elif cut == 'last':
            query = 'SELECT * FROM dtrs WHERE employee_id = %s AND YEAR(date) = %s AND MONTH(date) = %s AND DAY(date) >= 16'
            values = (employee_id, year, month)
        else:
            raise Exception('Wrong value')
        self.__database.cursor.execute(query, values)
        results = self.__database.cursor.fetchall()
        for result in results:
            temp = (result[2], result[3], result[4], result[5], result[6])
            dtrs.append(temp)
        return dtrs


    def __timedelta_to_time(self, delta:datetime.timedelta):
        fixed_date = datetime.datetime(2000, 1, 1)
        result_datetime = fixed_date + delta
        result_time = result_datetime.time()
        return result_time
    
    def verify_database(self):
        '''
        Verify if connected database is for municlock use
        '''
        admin_columns = ['id','username','password','name','level']
        dtr_columns = ['id','employee_id','date','am_in','am_out','pm_in','pm_out','locked']
        employees_columns = ['id','name','position','office','registered','am_in','am_out','pm_in','pm_out','noter','signatory','regular']
        officials_columns = ['id','name','position','signatory']
        biometrics_columns = ['id','name','ip_address','port','active']
        imports_columns = ['id','employee_id','created_at']
        logs_columns = ['id','admin_id','action','category','original','updated','created_at']
        noters_columns = ['id','name','position','office','signatory']

        query = f'''
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE 
            TABLE_NAME = "admins"
            AND TABLE_SCHEMA = "{self.__dbname}"
        '''
        self.__database.cursor.execute(query)
        admin_results = self.__database.cursor.fetchall()
        if admin_results:
            for column in admin_results:
                if column[0] not in admin_columns:
                    self.__dbverified = False
                    return False
        else: 
            return False
            
        query = f'''
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE 
            TABLE_NAME = "dtrs"
            AND TABLE_SCHEMA = "{self.__dbname}"
        '''
        self.__database.cursor.execute(query)
        dtr_results = self.__database.cursor.fetchall()
        if dtr_results:
            for column in dtr_results:
                if column[0] not in dtr_columns:
                    self.__dbverified = False
                    return False
        else:
            return False
        
        query = f'''
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE 
            TABLE_NAME = "employees"
            AND TABLE_SCHEMA = "{self.__dbname}"
        '''
        self.__database.cursor.execute(query)
        employee_results = self.__database.cursor.fetchall()
        if employee_results:
            for column in employee_results:
                if column[0] not in employees_columns:
                    self.__dbverified = False
                    return False
        else:
            return False
    
        query = f'''
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE 
            TABLE_NAME = "officials"
            AND TABLE_SCHEMA = "{self.__dbname}"
        '''
        self.__database.cursor.execute(query)
        official_results = self.__database.cursor.fetchall()
        if official_results:
            for column in official_results:
                if column[0] not in officials_columns:
                    self.__dbverified = False
                    return False
        else:
            return False
        
        query = f'''
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE 
            TABLE_NAME = "biometrics"
            AND TABLE_SCHEMA = "{self.__dbname}"
        '''
        self.__database.cursor.execute(query)
        biometric_results = self.__database.cursor.fetchall()
        if biometric_results:
            for column in biometric_results:
                if column[0] not in biometrics_columns:
                    self.__dbverified = False
                    return False
        else:
            return False
        
        query = f'''
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE 
            TABLE_NAME = "imports"
            AND TABLE_SCHEMA = "{self.__dbname}"
        '''
        self.__database.cursor.execute(query)
        import_results = self.__database.cursor.fetchall()
        if import_results:
            for column in import_results:
                if column[0] not in imports_columns:
                    self.__dbverified = False
                    return False
        else:
            return False

        query = f'''
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE 
            TABLE_NAME = "logs"
            AND TABLE_SCHEMA = "{self.__dbname}"
        '''
        self.__database.cursor.execute(query)
        log_results = self.__database.cursor.fetchall()
        if log_results:
            for column in log_results:
                if column[0] not in logs_columns:
                    self.__dbverified = False
                    return False
        else:
            return False
        
        query = f'''
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE 
            TABLE_NAME = "noters"
            AND TABLE_SCHEMA = "{self.__dbname}"
        '''
        self.__database.cursor.execute(query)
        noter_results = self.__database.cursor.fetchall()
        if noter_results:
            for column in noter_results:
                if column[0] not in noters_columns:
                    self.__dbverified = False
                    return False
        else:
            return False
        
        self.__dbverified = True
        return True
    
    def is_database_verified(self):
        '''
        Get if database is verified

        Returns `True` if database is verified, otherwise `False`
        '''
        return self.__dbverified

    def login_as_guest(self):
        '''
        Login as guest

        Grants level 1 access privilege
        '''
        self.__logged = True
        self.__id = 0
        self.__username = 'Guest'
        self.__name = 'Guest'
        self.__level = 1
        
    def login(self, username:str, password:str):
        '''
        Attempt to login. Current access level will also be \n
        set upon successful login 

        Returns `True` is successful, otherwise `False`
        '''
        info = self.__database.get_admin(username, password)
        if info:
            self.__logged = True
            self.__id = info[0]
            self.__username = info[1]
            self.__name = info[2]
            self.__level = info[3]
            return True
        else:
            return False
        
    def logout(self):
        '''
        Logout of current admin

        Returns `True` is successful, otherwise `False`
        '''
        if self.__logged:
            self.__logged = False
            self.__id = None
            self.username = ''
            self.name = ''
            self.__level = None
            return True
        else:
            return False

    def is_logged(self):
        '''
        Get current logged state

        Returns `True` if logged in, otherwise `False`
        '''
        return self.__logged
    
    def get_access_level(self):
        '''
        Get current access level

        Levels:
          - `1` : READ, EXPORT
          - `2` : READ, EXPORT, CREATE, UPDATE, DELETE
          - `3` : ALL
        '''
        return self.__level
    
    def get_username(self):
        '''
        Get current username
        '''
        return self.__username
    
    def get_name(self):
        '''
        Get current name
        '''
        return self.__name

    #################################################
    #                                               #
    #                    LEVEL 3                    #
    #                                               #
    #################################################

    def import_dtr(self, source:str, file:str = '', biometric:Biometric = None, start:datetime.date = None, end:datetime.date = None, on_failure_callback = None):
        '''
        Import a dtr-formatted text file or biometric\n
        Automatically cleaned and saved to database (imports table)

        Raises an Exception if access level is not high enough \n
        `source` - can be `file` or `biometric`\n
        `file` - file path (use if source is file)\n
        `biometric` - a biometric instance (use if source if biometric)\n
        `on_failure_callback` - function called on exception\n
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        self.busy = True

        dtrs = []

        # Reading file
        if source.lower() == 'file':
            origin = file
            file_type = file.split('.')[-1]
            if file_type.lower() == 'txt':
                with open(file) as dtr_file:
                    dtr_content = dtr_file.read()
                    rows =  dtr_content.split('\n')
                    rows.pop(0)
                    try:
                        int(rows[0].split()[2])
                        temp_datetime = datetime.datetime.strptime(f'{rows[0].split()[5]} {rows[0].split()[6]}', '%m/%d/%Y %H:%M:%S')
                    except Exception as e:
                        exception =  Exception('Wrong data format')
                        if on_failure_callback:
                            on_failure_callback(exception)
                        else:
                            raise Exception('Wrong data format')
                    for row in rows:
                        if row:
                            temp = row.split()
                            employee_number = temp[2]
                            date = temp[5]
                            time = temp[6]
                            temp_datetime = datetime.datetime.strptime(f'{date} {time}', '%m/%d/%Y %H:%M:%S')
                            temp_date = temp_datetime.date()
                            if start and end:
                                if not(temp_date >= start and temp_date <= end):
                                    continue
                            if start and not end:
                                if not(temp_date >= start):
                                    continue
                            if not start and end:
                                if not(temp_date <= end):
                                    continue
                            dtrs.append(f'{employee_number} {temp_date} {temp_datetime.time()}')
            elif file_type.lower() == 'xlsx':
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                try:
                    int(sheet['B1'].value)
                    temp_datetime = datetime.datetime.strptime(f'{sheet["A1"].value.date()} {sheet["D1"].value}', '%Y-%m-%d %H:%M:%S')
                except Exception as e:
                    if on_failure_callback:
                        exception =  Exception('Wrong data format')
                        on_failure_callback(exception)
                    else:
                        raise Exception('Wrong data format')
                for row in sheet.iter_rows():
                    employee_number = row[1].value
                    date = row[0].value.date()
                    time = row[3].value
                    temp_datetime = datetime.datetime.strptime(f'{date} {time}', '%Y-%m-%d %H:%M:%S')
                    temp_date = temp_datetime.date()
                    if start and end:
                        if not(temp_date >= start and temp_date <= end):
                            continue
                    if start and not end:
                        if not(temp_date >= start):
                            continue
                    if not start and end:
                        if not(temp_date <= end):
                            continue
                    dtrs.append(f'{employee_number} {temp_date} {temp_datetime.time()}')
            else:
                if on_failure_callback:
                    e = Exception('File type is not currently supported')
                    on_failure_callback(e)
                else:
                    raise Exception('File type is not currently supported')
        # Reading from biometric
        elif source.lower() == 'biometric':
            origin = biometric.ip
            try:
                attendances = biometric.get_attendance()
            except Exception as exception:
                if on_failure_callback:
                    on_failure_callback(exception)
            for attendance in attendances:
                temp_datetime = attendance.timestamp
                if start and end:
                    if not(temp_datetime.date() >= start and temp_datetime.date() <= end):
                        continue
                if start and not end:
                    if not(temp_datetime.date() >= start ):
                        continue
                if not start and end:
                    if not(temp_datetime.date() <= end):
                        continue
                employee_number = attendance.user_id
                dtrs.append(f'{employee_number} {temp_datetime.date()} {temp_datetime.time()}')

        # Removing duplicated minutes
        dtrs = set(dtrs)

        # Sorting DTRs by datetime
        sorted_dtrs = sorted(dtrs, key = lambda item: datetime.datetime.strptime(f'{item.split()[1]} {item.split()[2]}', '%Y-%m-%d %H:%M:%S'))

        for dtr in sorted_dtrs:
            temp = dtr.split(' ')
            employee_id = temp[0]
            created_at = f'{temp[1]} {temp[2]}'
            self.__database.add_import(employee_id, created_at, commit = False)

        self.__database.commit()
        self.busy = False

        self.__database.add_log(
            admin_id = self.__id,
            action = 'import',
            category = 'import',
            original = f'',
            updated = f'import(source = {source}, origin = {origin})'
        )
    
    def admin_exists(self, id:str|int):
        '''
        Check if an admin exists by id

        Returns `True` if exists, otherwise `False`
        '''
        return self.__database.admin_exists(id)
    
    def add_admin(self, username:str, password:str, name:str, level:int|str):
        '''
        Add admin account

        `password` should be a raw string as passwords are automatically hashed when saved to database
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        self.__database.add_log(
            admin_id = self.__id, 
            action = 'add', 
            category = 'admin',
            original = '', 
            updated = f'admin(username = {username}, password = hidden, name = {name}, level = {level})')
        return self.__database.add_admin(username, password, name, level)
    
    def get_admins(self, order_by:str = 'id', order:str = 'asc'):
        '''
        Get all admin accounts

        Returns a list of tupples \n
        `order_by` can be `id`, `name`, `username`, `level` \n
        `order` can be `asc` or `desc`, default to `asc`
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_admins(order_by = order_by, order = order)
    
    def get_admin(self, id:str|int):
        '''
        Get admin account by id
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_admin_by_id(id)
    
    def update_admin_info(self, original_id:str|int, new_id:str|int, username:str, name:str, level:int|str, commit:bool = True):
        '''
        Update an admin account info

        Password is not included \n
        Use `reset_admin_password` instead \n

        Returns `True` if successfully updated, otherwise `False`. \n
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        original_admin = self.__database.get_admin_by_id(original_id)
        if self.__database.update_admin_info(original_id, new_id, username, name, level, commit = commit):
            self.__database.add_log(
                admin_id = self.__id, 
                action = 'update', 
                category = 'admin',
                original = f'admin(id = {original_admin[0]}, username = {original_admin[1]}, name = {original_admin[3]}, level = {original_admin[4]})',
                updated = f'admin(id = {new_id}, username = {username}, name = {name}, level = {level})'
            )
            return True
        else:
            return False
    
    def reset_admin_password(self, id:str|int, password:str, commit:bool = True):
        '''
        Reset admin password

        Returns `True` if successfully updated, otherwise `False`. \n
        Use with caution
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        if self.__database.reset_admin_password(id, password, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'update',
                category = 'admin',
                original = f'admin(password = hidden)',
                updated = f'admin(password = hidden)'
            )
            return True
        else:
            return False

    def delete_admin(self, id:str|int, commit:bool = True):
        '''
        Delete admin account

        Return `True` is successfully deleted, otherwise `False` if account does not exists \n
        If `commit` is True, automatically commit after updating. Default to `True`
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        original_admin = self.__database.get_admin_by_id(id)
        if self.__database.delete_admin(id, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'delete',
                category = 'admin',
                original = f'admin(id = {id}, username = {original_admin[1]}, name = {original_admin[3]}, level = {original_admin[4]})',
                updated = f''
            )
            return True
        else:
            return False

    def add_biometric(self, name:str, ip_address:str, port:str|int, active:bool|int = False, commit:bool = True):
        '''
        Add a biometric

        Return `True` is successfully add \n
        If `commit` is True, automatically commit after adding. Default to `True`
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        if self.__database.add_biometric(name, ip_address, port, active = active, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'add',
                category = 'biometric',
                original = f'',
                updated = f'biometric(name = {name}, ip = {ip_address}, port = {port}, active = {int(active)})'
            )
            return True
        else:
            return False
    
    def biometric_exists(self, id:str|int):
        '''
        Check if a biometric exists

        Returns `True` if exists otherwise returns `False`
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.biometric_exists(id = id)
    
    def biometric_ip_exists(self, ip_address:str):
        '''
        Check if a biometric IP address exists

        Returns `True` if exists otherwise returns `False`
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.biometric_ip_exists(ip_address = ip_address)

    def get_biometrics(self, order_by:str = 'id', order:str = 'asc'):
        '''
        Get all biometrics model

        Returns a list of tupples
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_biometrics(order_by = order_by, order = order)
    
    def get_biometric(self, id:str|int):
        '''
        Get a biometric by id
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_biometric(id = id)
    
    def get_biometric_by_ip(self, ip_address:str):
        '''
        Get a biometric by ip address
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_biometric_by_ip(ip_address = ip_address)
    
    def update_biometric(self, id:str|int, new_name:str = '', new_ip:str = '', new_port:str = '', new_active:bool|int = False, commit:bool = True):
        '''
        Update a biometric by id

        Return `True` is successfully updated, otherwise `False` if account does not exists \n
        If `commit` is True, automatically commit after updating. Default to `True`
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        original_biometric = self.__database.get_biometric(id)
        if self.__database.update_biometric(id = id, new_name = new_name, new_ip = new_ip, new_port = new_port, new_active = new_active, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'update',
                category = 'biometric',
                original = f'biometric(name = {original_biometric[1]}, ip = {original_biometric[2]}, port = {original_biometric[3]}, active = {original_biometric[4]})',
                updated = f'biometric(name = {new_name}, ip = {new_ip}, port = {new_port}, active = {int(new_active)})'
            )
            return True
        else:
            return False
    
    def delete_biometric(self, id:str|int, commit:bool = True):
        '''
        Delete a biometric

        Return `True` is successfully deleted, otherwise `False` if account does not exists \n
        If `commit` is True, automatically commit after deleting. Default to `True`
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        original_biometric = self.__database.get_biometric(id)
        if self.__database.delete_biometric(id = id, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'delete',
                category = 'biometric',
                original = f'biometric(id = {original_biometric[0]}, name = {original_biometric[1]}, ip = {original_biometric[2]}, port = {original_biometric[3]}, active = {original_biometric[4]})',
                updated = f''
            )
    
    def test_biometric_connection(self, ip:str, port:int, timeout:int = 10):
        '''
        Test connection to a biometric

        `timeout` can be set for the connection test timeout \n
        Returns `True` if test succeed, otherwise returns `False`
        '''
        temp_biometric = zk.ZK(ip = ip, port = port, timeout = timeout, password = 0, force_udp = False, ommit_ping = False)

        try:
            temp_biometric.connect()
            return True
        except:
            return False

    def connect_to_biometric(self, ip:str, port:str|int, timeout:int = 1000, name:str = '', test_timeout = 10):
        '''
        Connect to a biometric.

        Connection is first tested and returns a Biometric object on successful connection\n
        Returns `None` if connection was unsuccessful
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')

        if not self.test_biometric_connection(ip, port, test_timeout):
            return None
        
        biometric = Biometric(ip = ip, port = port, timeout = timeout, name = name)
        return biometric
    
    def activate_biometric(self, id:str|int):
        '''
        Activate a biometric

        Return `True` is successfully activated, otherwise `False` if biometric does not exists \n
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        if self.__database.activate_biometric(id = id):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'update',
                category = 'biometric',
                original = f'',
                updated = f'biometric(id = {id}, active = 1)'
            )
            return True
        else:
            return False
    
    def deactivate_biometric(self, id:str|int):
        '''
        Activate a biometric

        Return `True` is successfully deactivated, otherwise `False` if biometric does not exists \n
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        if self.__database.deactivate_biometric(id = id):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'update',
                category = 'biometric',
                original = f'',
                updated = f'biometric(id = {id}, active = 0)'
            )
            return True
        else:
            return False

    def get_active_biometrics(self):
        '''
        Get currently active biometrics
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_active_biometrics()

    def import_exists(self, employee_id:str|int, created_at:str|datetime.datetime):
        '''
        Check if an import exists

        Returns `True` if exists otherwise returns `False`
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.import_exists(employee_id, created_at)
    
    def add_import(self, employee_id:str|int, created_at:str|datetime.datetime, commit:bool = True):
        '''
        Add an import to import table

        Return `True` is successfully add \n
        Returns `False` if created_at and employee_id exists \n
        If `commit` is True, automatically commit after adding. Default to `True`
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        if self.__database.add_import(employee_id, created_at, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'add',
                category = 'import',
                original = f'',
                updated = f'import(employee_id = {employee_id}, created_at = {created_at})'
            )
            return True
        else:
            return False
    
    def get_imports(self, employee_ids:list = [], start:str|datetime.date = '', end:str|datetime.date = ''):
        '''
        Get imports by ids and select date range

        `employee_ids` - a list of valid employee ids to find. If empty returns get all employee_id. Default to empty \n
        Returns an exception if arguments are invalid\n
        `start` - start range (optional) \n
        `end` - end range (optional)
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_imports(employee_ids = employee_ids, start = start, end = end)

    def get_logs(self, start:datetime.date|str = None, end:datetime.date|str = None):
        '''
        Get logs

        If `start` and `end` is omitted, returns all logs \n
        If `start` is only defined, returns logs from that point forward \n
        If `end` is only defined, returns logs until that point forward \n
        '''
        if self.__level is None or self.__level < 3:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_logs(start = start, end = end)
    
    def generate_logs(self, start:datetime.date|str = None, end:datetime.date|str = None):
        logs = self.get_logs(start = start, end = end)
        content = ''
        header = f'Logs as of {datetime.datetime.now()}\n\n'
        content += header
        for log in logs:
            message = f'''{log[5]}
    Admin Name:{log[0]}
    Action:    {log[1]} 
    Field:     {log[2]} 
    Original:  {log[3].ljust(100, " ")} 
    Updated:   {log[4].ljust(100, " ")}
'''
            content += message

        if not os.path.isdir(f'{self.export_folder}\logs'):
            os.makedirs(f'{self.export_folder}\logs')

        with open(f'{self.export_folder}/logs/{datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")}.log', 'w') as file:
            file.write(content)
    
    #################################################
    #                                               #
    #                    LEVEL 2                    #
    #                                               #
    #################################################

    def refresh_dtr(self, employee_id:int|str = 0):
        '''
        Refresh employees DTRs\n
        Data is retrieved from the import table

        If `employee_id` is supplied refresh dtr of specific employee instead
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        try:
            employee_id = int(employee_id)
        except:
            raise Exception('Employee is not whole number')
        
        self.busy = True
        employees = self.__database.get_employees()
        
        if employee_id == 0:
            employees_ids = []
            for employee in employees:
                employees_ids.append(employee[0])
            dtrs = self.__database.get_imports(employee_ids = employees_ids)
        else:
            dtrs = self.__database.get_imports(employee_ids = [employee_id])

        sorted_data = collections.defaultdict(lambda: collections.defaultdict(set))

        # Group the data by employee_id and date
        for dtr in dtrs:
            id_val, emp_id, dt = dtr
            date_str = dt.date()
            time_str = dt.time()
            sorted_data[emp_id][date_str].add(time_str)

        # Sort the data and format the output
        formatted_data = {}
        for emp_id, date_data in sorted_data.items():
            emp_data = {}
            for date, times in date_data.items():
                emp_data[str(date)] = sorted(times)
            formatted_data[emp_id] = emp_data

        for employee_id, employee_data in formatted_data.items():
            regular_am_in = self.get_employee_am_in(employee_id)
            regular_am_out = self.get_employee_am_out(employee_id)
            regular_pm_in = self.get_employee_pm_in(employee_id)
            regular_pm_out = self.get_employee_pm_out(employee_id)
            for date, times in employee_data.items():
                am_in = ''
                am_out = ''
                pm_in = ''
                pm_out = ''
                sorted_times = sorted(times)
                for index, time in enumerate(sorted_times):
                    hours = time.hour
                    minutes = time.minute
                    seconds = time.second
                    time_object = datetime.timedelta(hours = hours, minutes = minutes, seconds = seconds)

                    # Check for possible afternoon halfday
                    if len(sorted_times) == 2 and index < len(sorted_times) - 1:
                        next_hours = sorted_times[index + 1].hour
                        next_minutes = sorted_times[index + 1].minute
                        next_time_object = datetime.timedelta(hours = next_hours, minutes = next_minutes)
                    else:
                        next_time_object = datetime.timedelta(hours = 0, minutes = 0)

                    # Assign times to respective check in and out
                    if not am_in and time_object <= regular_am_out:
                        am_in = time_object
                        continue
                    if not am_out and time_object >= regular_am_in and time_object <= regular_pm_in:
                        # Check if next time object exist (can only have value if possible afternoon halfday)
                        if not next_time_object >= regular_pm_in:
                            if am_in:
                                if (time_object - am_in) > datetime.timedelta(minutes = 1, seconds = 0):
                                    am_out = time_object
                            else:
                                am_out = time_object
                            continue
                    if not pm_in and time_object >= regular_am_out and time_object <= regular_pm_out - datetime.timedelta(hours = 1):
                        if am_out:
                            if(time_object - am_out) > datetime.timedelta(minutes = 1, seconds = 0):
                                pm_in = time_object
                        else:
                            pm_in = time_object
                        continue
                    if not pm_out and time_object > regular_pm_in:
                        if pm_in:
                            if(time_object - pm_in) >= datetime.timedelta(minutes = 1, seconds = 0):
                                pm_out = time_object
                        else:
                            pm_out = time_object
                        continue
                
                # Convert timedelta to time object
                if am_in:
                    am_in = self.__timedelta_to_time(am_in)
                if am_out:
                    am_out = self.__timedelta_to_time(am_out)
                if pm_in:
                    pm_in = self.__timedelta_to_time(pm_in)
                if pm_out:
                    pm_out = self.__timedelta_to_time(pm_out)

                # Check if a dtr with current date exists
                if self.__database.check_dtr_date_exists(employee_id, date):
                    dtr_by_date = self.get_employee_dtr_by_date(employee_id, date)
                    if dtr_by_date[7] == 0:
                        self.__database.update_dtr(dtr_by_date[0], date, am_in = am_in, am_out = am_out, pm_in = pm_in, pm_out = pm_out, locked = 0, commit = False)
                else:
                    self.__database.add_dtr(employee_id, date, am_in = am_in, am_out = am_out, pm_in = pm_in, pm_out = pm_out, locked = 0, commit = False)

        self.__database.commit()

        self.busy = False

    def get_official(self, id:int|str):
        '''
        Get employee data

        Returns a tupple containing employee data \n
        (id, name, position, office, registered, time_in, time_out) \n
        Returns `None` if not found
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_official(id)
    
    def add_official(self, name:str, position:str, signatory:str, commit:bool = True):
        '''
        Add employee to database

        Returns `True` if successfully saved, otherwise `False`. 
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        if self.__database.add_official(name, position, signatory, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'add',
                category = 'official',
                original = f'',
                updated = f'official(name = {name}, position = {position}, signatory = {signatory})'
            )
            return True
        else:
            return False
    
    def update_official(self, id:int|str, new_name:str, new_position:str, new_signatory:str, commit:bool = True):
        '''
        Update an official by id

        Returns `True` if successfully updated, otherwise `False` \n
        Returns `False` if new_id supplied is already assigned to an employee \n
        If `commit` is True, automatically commit after adding. Default to `True`
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        original_official = self.__database.get_official(id)
        if self.__database.update_official(id, new_name, new_position, new_signatory, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'update',
                category = 'official',
                original = f'official(id = {original_official[0]}, name = {original_official[1]}, position = {original_official[2]}, signatory = {original_official[3]})',
                updated = f'official(name = {new_name}, position = {new_position}, signatory = {new_signatory})'
            )
            return True
        else:
            return False

    def delete_official(self, id:int|str, commit:bool = True):
        '''
        Delete selected official
        
        Returns `True` if successfully updated, otherwise `False` \n
        If `commit` is True, automatically commit after updating. Default to `True`
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        original_official = self.__database.get_official(id)
        if self.__database.delete_official(id, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'delete',
                category = 'official',
                original = f'official(id = {id}, name = {original_official[1]}, position = {original_official[2]}, signatory = {original_official[3]})',
                updated = f''
            )
            return True
        else:
            return False

    def add_employee(self, id:int|str, name:str, position:str, office:str, registered:bool|int, am_in:str|datetime.time, am_out:str|datetime.time, pm_in:str|datetime.time, pm_out:str|datetime.time, noter:bool|int, signatory:str, regular:bool|int, commit:bool = True):
        '''
        Add employee to database

        Returns `True` if successfully saved, otherwise `False`. \n
        Also returns `False` if duplicated id\n
        Returns an exception if args are not valid
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        if self.__database.add_employee(id, name, position, office, registered, am_in, am_out, pm_in, pm_out, noter, signatory, regular, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'add',
                category = 'employee',
                original = f'',
                updated = f'employee(id = {id}, name = {name}, position = {position}, office = {office}, registered = {int(registered)}, am_in = {am_in}, am_out = {am_out}, pm_in = {pm_in}, pm_out = {pm_out}, noter = {int(noter)}, signatory = {signatory}, regular = {int(regular)})'
            )
            return True
        else:
            return False
    
    def update_employee(self, original_id: int|str, new_id:int|str, new_name:str, new_position:str, new_office:str, new_registered:bool|int, new_am_in:str|datetime.time, new_am_out:str|datetime.time, new_pm_in:str|datetime.time, new_pm_out:str|datetime.time, new_noter:bool|int, new_signatory:str, new_regular:bool|int, commit:bool = True):
        '''
        Update an employee by id

        Returns `True` if successfully updated, otherwise `False` \n
        Returns `False` if new_id supplied is already assigned to an employee \n
        If `commit` is True, automatically commit after adding. Default to `True`\n
        Returns an exception if args are not valid
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        original_employee = self.get_employee(original_id)
        if self.__database.update_employee(original_id, new_id, new_name, new_position, new_office, new_registered, new_am_in, new_am_out, new_pm_in, new_pm_out, new_noter, new_signatory, new_regular, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'update',
                category = 'employee',
                original = f'employee(id = {original_id}, name = {original_employee[1]}, position = {original_employee[2]}, office = {original_employee[3]}, registered = {original_employee[4]}, am_in = {original_employee[5]}, am_out = {original_employee[6]}, pm_in = {original_employee[7]}, pm_out = {original_employee[8]}, noter = {original_employee[9]}, signatory = {original_employee[10]}, regular = {original_employee[11]})',
                updated = f'employee(id = {new_id}, name = {new_name}, position = {new_position}, office = {new_office}, registered = {int(new_registered)}, am_in = {new_am_in}, am_out = {new_am_out}, pm_in = {new_pm_in}, pm_out = {new_pm_out}, noter = {int(new_noter)}, signatory = {new_signatory}, regular = {int(new_regular)})'
            )
            return True
        else:
            return False
    
    def delete_employee(self, id:int|str, commit:bool = True):
        '''
        Delete selected employee
        
        Returns `True` if successfully updated, otherwise `False` \n
        If `commit` is True, automatically commit after updating. Default to `True`
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        original_employee = self.get_employee(id)
        if self.__database.delete_employee(id, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'delete',
                category = 'employee',
                original = f'employee(id = {id}, name = {original_employee[1]}, position = {original_employee[2]}, office = {original_employee[3]}, registered = {original_employee[4]}, am_in = {original_employee[5]}, am_out = {original_employee[6]}, pm_in = {original_employee[7]}, pm_out = {original_employee[8]}, noter = {original_employee[9]}, signatory = {original_employee[10]}, regular = {original_employee[11]})',
                updated = f''
            )
            return True
        else:
            return False

    def add_dtr(self, employee_id:int|str, date:str|datetime.date, am_in:str|datetime.time = '', am_out:str|datetime.time = '', pm_in:str|datetime.time = '', pm_out:str|datetime.time = '', locked:bool|int = 0, commit:bool = True):
        '''
        Add DTR to an employee

        Returns `True` if successfully added, otherwise `False` \n
        Also returns `False` if date exist \n
        If times are omitted or empty, null will be saved on respective column \n
        If `commit` is True, automatically commit after adding. Default to `True`\n
        Returns an exception if arguments are not valid
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        if self.__database.add_dtr(employee_id, date, am_in = am_in, am_out = am_out, pm_in = pm_in, pm_out = pm_out, locked = locked, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'add',
                category = 'dtr',
                original = f'',
                updated = f'dtr(employee_id = {employee_id}, date = {date}, am_in = {am_in}, am_out = {am_out}, pm_in = {pm_in}, pm_out = {pm_out}, locked = {int(locked)})'
            )
            return True
        else:
            return False
    
    def update_dtr(self, id:int|str, date:str|datetime.date, am_in:str|datetime.time = '', am_out:str|datetime.time = '', pm_in:str|datetime.time = '', pm_out:str|datetime.time = '', locked:bool|int = 0, commit:bool = True):
        '''
        Update a DTR given by id

        Returns `True` if successfully updated, otherwise `False` \n
        Also returns `False` if date exist \n
        If times are omitted or empty, null will be saved on respective column \n
        If `commit` is True, automatically commit after updating. Default to `True`\n
        Returns an exception if arguments are not valid
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        original_dtr = self.get_dtr(id)
        if self.__database.update_dtr(id, date, am_in = am_in, am_out = am_out, pm_in = pm_in, pm_out = pm_out, locked = locked, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'update',
                category = 'dtr',
                original = f'dtr(employee_id = {id}, date = {original_dtr[2]}, am_in = {original_dtr[3]}, am_out = {original_dtr[4]}, pm_in = {original_dtr[5]}, pm_out = {original_dtr[6]}, locked = {int(original_dtr[7])})',
                updated = f'dtr(employee_id = {id}, date = {date}, am_in = {am_in}, am_out = {am_out}, pm_in = {pm_in}, pm_out = {pm_out}, locked = {int(locked)})'
            )
            return True
        else:
            return False
    
    def delete_dtr(self, id:int|str, commit:bool = True):
        '''
        Delete a DTR by id

        Returns `True` if successfully updated, otherwise `False`
        If `commit` is True, automatically commit after updating. Default to `True`
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        original_dtr = self.get_dtr(id)
        if self.__database.delete_dtr(id, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'delete',
                category = 'dtr',
                original = f'dtr(employee_id = {id}, date = {original_dtr[2]}, am_in = {original_dtr[3]}, am_out = {original_dtr[4]}, pm_in = {original_dtr[5]}, pm_out = {original_dtr[6]}, locked = {original_dtr[7]})',
                updated = f''
            )
            return True
        else:
            return False

    def unimported_dtr_exists(self, employee_id:int|str):
        '''
        Check if an employee dtr exists on imports table but not exists in dtrs

        Returns `True` if unimported dtr exists, otherwise returns `False`
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        return self.__database.unimported_dtr_exists(employee_id)

    def add_noter(self, name:str, position:str, office:str, signatory:str, commit:bool = True):
        '''
        Add noter to database

        Returns `True` if successfully saved, otherwise `False`. \n
        Also returns `False` if duplicated id\n
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')

        if self.__database.add_noter(name, position, office, signatory, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'add',
                category = 'noter',
                original = f'',
                updated = f'noter(name = {name}, position = {position}, office = {office}, signatory = {signatory})'
            )
            return True
        else:
            return False
    
    def update_noter(self, id: int|str, new_name:str, new_position:str, new_office:str, new_signatory:str, commit:bool = True):
        '''
        Update a noter by id

        Returns `True` if successfully updated, otherwise `False` \n
        Returns `False` if noter does not exist \n
        If `commit` is True, automatically commit after adding. Default to `True`\n
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        original_noter = self.get_noter(id)
        if self.__database.update_noter(id, new_name, new_position, new_office, new_signatory, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'update',
                category = 'noter',
                original = f'noter(id = {original_noter[0]}, name = {original_noter[1]}, position = {original_noter[2]}, office = {original_noter[3]}, signatory = {original_noter[4]})',
                updated = f'noter(id = {original_noter[0]}, name = {new_name}, position = {new_position}, office = {new_office}, signatory = {new_signatory})'
            )
            return True
        else:
            return False
    
    def delete_noter(self, id: int|str, commit:bool = True):
        '''
        Delete a noter by id

        Returns `True` if successfully deleted, otherwise `False` \n
        Returns `False` if noter does not exist \n
        If `commit` is True, automatically commit after adding. Default to `True`
        '''
        if self.__level is None or self.__level < 2:
            raise Exception('Insufficient access privileges')
        
        original_noter = self.get_noter(id)
        if self.__database.delete_noter(id, commit = commit):
            self.__database.add_log(
                admin_id = self.__id,
                action = 'delete',
                category = 'noter',
                original = f'noter(id = {original_noter[0]}, name = {original_noter[1]}, position = {original_noter[2]}, office = {original_noter[3]}, signatory = {original_noter[4]})',
                updated = f''
            )
            return True
        else:
            return False

    #################################################
    #                                               #
    #                    LEVEL 1                    #
    #                                               #
    #################################################

    def export_dtr(self, employee_id:int|str, noter_signatory:str, noter_position:str, first_month:int|str, first_year:int|str, first_cut:str = 'full', second_month:int|str = 0, second_year:int|str = 0, second_cut:str = 'full', export_to:str = 'excel', preview:bool = False):
        '''
        Export an employee DTR on a certain month and year \n
        `second_month` and `second_year` can be null \n
        `noter_id` - employee_id of noter (Set to 0 to blank) \n
        Defaults to excel exports

        Supported exports:
          - `excel`
          - `pdf`
        '''
        self.busy = True

        if second_year == 0:
            second_year = first_year

        months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        first_month = int(first_month)
        first_year = int(first_year)
        second_month = int(second_month)
        second_year = int(second_year)

        employee = self.__database.get_employee(employee_id)
        employee_name = employee[1]
        employee_signatory = employee[10]
        employee_position = employee[2]
        employee_office = employee[3]
        employee_in = employee[5]
        employee_out = employee[8]

        if first_cut.lower() == 'full':
            first_date_range_string = f'1 - {calendar.monthrange(first_year, first_month)[1]}'
        elif first_cut.lower() == 'first':
            first_date_range_string = '1 - 15'
        elif first_cut.lower() == 'last':
            first_date_range_string = f'16 - {calendar.monthrange(first_year, first_month)[1]}'

        if second_month != 0 and second_year != 0:
            if second_cut.lower() == 'full':
                second_date_range_string = f'1 - {calendar.monthrange(second_year, second_month)[1]}'
            elif second_cut.lower() == 'first':
                second_date_range_string = '1 - 15'
            elif second_cut.lower() == 'last':
                second_date_range_string = f'16 - {calendar.monthrange(second_year, second_month)[1]}'

        in_total_seconds = employee_in.total_seconds()
        in_hours, in_remainder = divmod(in_total_seconds, 3600)
        in_minutes, _ = divmod(in_remainder, 60)
        in_meridian = 'AM' if in_hours < 12 else 'PM'
        in_formatted_hour = int(in_hours) % 12
        if in_formatted_hour == 0:
            in_formatted_hour = 12
        in_formatted_time = f"{in_formatted_hour:02d}:{int(in_minutes):02d}{in_meridian}"

        out_total_seconds = employee_out.total_seconds()
        out_hours, out_remainder = divmod(out_total_seconds, 3600)
        out_minutes, _ = divmod(out_remainder, 60)
        out_meridian = 'AM' if out_hours < 12 else 'PM'
        out_formatted_hour = int(out_hours) % 12
        if out_formatted_hour == 0:
            out_formatted_hour = 12
        out_formatted_time = f"{out_formatted_hour:02d}:{int(out_minutes):02d}{out_meridian}"

        regular_time = f'{in_formatted_time} - {out_formatted_time}'

        now = datetime.datetime.now()
        if preview:
            previews_dir = os.path.join(self.export_folder, 'previews')
            os.makedirs(previews_dir, exist_ok=True)
            filename = os.path.join(previews_dir, str(employee_id))
        else:
            office_name_dir = os.path.join(self.export_folder, employee_office, employee_name)
            os.makedirs(office_name_dir, exist_ok=True)
            second_name_string = ''
            if second_month != 0:
                if months[first_month - 1] != months[second_month - 1]:
                        second_name_string = f' - {months[second_month - 1]} {second_year}'
            filename = os.path.join(office_name_dir, f'{months[first_month - 1]} {first_year} {second_name_string} v({now.date()}-{now.hour}-{now.minute}-{now.second})')

        if export_to.lower() == 'excel':
            workbook = openpyxl.load_workbook(filename = f'{self._directory}/templates/format.xlsx')
            sheet = workbook.active

            # Filling in first DTR Excel Template
            sheet['A4'] = employee_name
            sheet['A6'] = f'{months[first_month - 1]} {first_date_range_string}, {first_year}'
            sheet['F7'] = regular_time
            sheet['C48'] = employee_signatory
            sheet['C49'] = employee_position
            sheet['C52'] = noter_signatory
            sheet['C53'] = noter_position

            first_employee_dtr = self.__get_dtr(employee_id, first_month, first_year, first_cut)
            for dtr in first_employee_dtr:
                row = dtr[0].day + 10
                for i in range(1, len(dtr)):
                    if i == 1:
                        column = 'B'
                    elif i == 2:
                        column = 'C'
                    elif i == 3:
                        column = 'D'
                    elif i == 4:
                        column = 'E'
                    if dtr[i]:
                        total_seconds = dtr[i].total_seconds()
                        hours, remainder = divmod(total_seconds, 3600)
                        minutes, _ = divmod(remainder, 60)
                        if hours > 12:
                            hours -= 12
                        sheet[f'{column}{row}'] = f'{int(hours):02d}:{int(minutes):02d}'

            # Filling in second DTR Excel Template
            if second_month != 0 and second_year != 0:
                sheet['I4'] = employee_name
                sheet['I6'] = f'{months[second_month - 1]} {second_date_range_string}, {second_year}'
                sheet['N7'] = regular_time
                sheet['K48'] = employee_signatory
                sheet['K49'] = employee_position
                sheet['K52'] = noter_signatory
                sheet['K53'] = noter_position
                
                second_employee_dtr = self.__get_dtr(employee_id, second_month, second_year, second_cut)
                for dtr in second_employee_dtr:
                    row = dtr[0].day + 10
                    for i in range(1, len(dtr)):
                        if i == 1:
                            column = 'J'
                        elif i == 2:
                            column = 'K'
                        elif i == 3:
                            column = 'L'
                        elif i == 4:
                            column = 'M'
                        if dtr[i]:
                            total_seconds = dtr[i].total_seconds()
                            hours, remainder = divmod(total_seconds, 3600)
                            minutes, _ = divmod(remainder, 60)
                            if hours > 12:
                                hours -= 12
                            sheet[f'{column}{row}'] = f'{int(hours):02d}:{int(minutes):02d}'

            workbook.save(filename = filename+'.xlsx')
        
        elif export_to.lower() == 'pdf':
            packet = io.BytesIO()
            can = canvas.Canvas(packet, pagesize=letter)
            pdfmetrics.registerFont(TTFont('Calibri', f'{self._directory}/templates/Calibri.ttf'))
            pdfmetrics.registerFont(TTFont('Calibri Bold', f'{self._directory}/templates/Calibri Bold.ttf'))
            pdfmetrics.registerFont(TTFont('Segoe UI', f'{self._directory}/templates/Segoe UI.ttf'))
            pdfmetrics.registerFont(TTFont('Segoe UI Bold', f'{self._directory}/templates/Segoe UI Bold.ttf'))
            pdfmetrics.registerFont(TTFont('Times New Roman', f'{self._directory}/templates/Times New Roman.ttf'))

            # First Page
            can.setFont('Segoe UI Bold', 10)
            can.drawCentredString(168.5, 740, employee_name)

            can.setFont('Calibri Bold', 9)
            can.drawCentredString(168.5, 717.5, f'{months[first_month - 1]} {first_date_range_string}, {first_year}')

            can.setFont('Times New Roman', 8)
            can.drawString(213, 704, regular_time)

            can.setFont('Segoe UI Bold', 9)
            can.drawCentredString(178.5, 213.5, employee_signatory)

            can.setFont('Times New Roman', 9)
            can.drawCentredString(178.5, 201, employee_position)

            can.setFont('Segoe UI Bold', 9)
            can.drawCentredString(178.5, 164, noter_signatory)

            can.setFont('Times New Roman', 9)
            can.drawCentredString(178.5, 152, noter_position)

            # time data
            first_employee_dtr = self.__get_dtr(employee_id, first_month, first_year, first_cut)
            can.setFont('Calibri', 9)

            for dtr in first_employee_dtr:
                y = 656.5 - ((dtr[0].day - 1) * 12)
                for i in range(1, len(dtr)):
                    if i == 1:
                        x = 78.5
                    elif i == 2:
                        x = 114.2
                    elif i == 3:
                        x = 149.9
                    elif i == 4:
                        x = 185.6
                    if dtr[i]:
                        total_seconds = dtr[i].total_seconds()
                        hours, remainder = divmod(total_seconds, 3600)
                        minutes, _ = divmod(remainder, 60)
                        if hours > 12:
                            hours -= 12
                        can.drawString(x, y, f'{int(hours):02d}:{int(minutes):02d}')
                        
            # Second Page
            if second_month != 0 and second_year != 0:
                can.setFont('Segoe UI Bold', 10)
                can.drawCentredString(423, 740, employee_name)

                can.setFont('Calibri Bold', 9)
                can.drawCentredString(423, 717.5, f'{months[second_month - 1]} {second_date_range_string}, {second_year}')

                can.setFont('Times New Roman', 8)
                can.drawString(467.5, 704, regular_time)

                can.setFont('Segoe UI Bold', 9)
                can.drawCentredString(433, 213.5, employee_signatory)

                can.setFont('Times New Roman', 9)
                can.drawCentredString(433, 201, employee_position)

                can.setFont('Segoe UI Bold', 9)
                can.drawCentredString(433, 164, noter_signatory)

                can.setFont('Times New Roman', 9)
                can.drawCentredString(433, 152, noter_position)

                # time data
                can.setFont('Calibri', 9)
                second_employee_dtr = self.__get_dtr(employee_id, second_month, second_year, second_cut)
                can.setFont('Calibri', 9)

                for dtr in second_employee_dtr:
                    y = 656.5 - ((dtr[0].day - 1) * 12)
                    for i in range(1, len(dtr)):
                        if i == 1:
                            x = 333
                        elif i == 2:
                            x = 368.7
                        elif i == 3:
                            x = 404.4
                        elif i == 4:
                            x = 440.1
                        if dtr[i]:
                            total_seconds = dtr[i].total_seconds()
                            hours, remainder = divmod(total_seconds, 3600)
                            minutes, _ = divmod(remainder, 60)
                            if hours > 12:
                                hours -= 12
                            can.drawString(x, y, f'{int(hours):02d}:{int(minutes):02d}')
            can.save()

            packet.seek(0)
            new_pdf = PdfReader(packet)
            template_pdf = PdfReader(open(f'{self._directory}/templates/format.pdf', 'rb'))
            output = PdfWriter()
            page = template_pdf.pages[0]
            page.merge_page(new_pdf.pages[0])
            output.add_page(page)
            output_stream = open(f'{filename}.pdf', 'wb')
            output.write(output_stream)
            output_stream.close()

        self.busy = False

    def noter_exists(self, noter_id:int|str):
        '''
        Check if an noter exist
        
        Returns `true` if exist in database else `false`
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.noter_exists(noter_id)

    def get_noter(self, noter_id:int|str):
        '''
        Get a noter by id
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_noter(noter_id)
    
    def get_noters(self, include_others:bool = False, order_by:str = 'id', order:str = 'asc'):
        '''
        Get all employees, officials, and noters that is qualified to sign on DTR\n
        If `include_others` is True, includes officials and employees to the list.\n
        Ordering is only applicable if `include_others` is False

        Returns a list of tupples
        if `include_others` is True: (signatory, position)
        else: (id, name, position, office, signatory)
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_noters(include_others = include_others, order_by = order_by, order = order)
    
    def get_officials(self, order_by:str = 'id', order:str = 'asc'):
        '''
        Get all officials

        Returns a list of tupples \n
        `order_by` can be `id`, `name`, `position` \n
        `order` can be `asc` or `desc`, default to `asc`
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_officials(order_by = order_by, order = order)

    def get_official_positions(self, order:str = 'asc'):
        '''
        Get all distinct official positions

        Returns a list of all distinct positions \n
        `order` can be `asc` or `desc`, default to `asc`. If order is not valid returns an emply list 
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_official_positions(order = order)
    
    def get_employees(self, employee_type:str = 'all', order_by:str = '', order:str = 'asc', limit:int = 1000, offset:int = 0):
        '''
        Get all employees

        Returns a list of tupples \n
        `order_by` can be `id`, `name`, `position`, `office`, `registered`, `time_in`, `time_out` \n
        `order` can be `asc` or `desc`, default to `asc`
        `employee_type` - can be `all`, `regular`, `job order`. Default to all.
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employees(employee_type = employee_type, order_by = order_by, order = order, limit = limit, offset = offset)

    def get_employees_by_office(self, office:str, employee_type:str = 'all', order_by:str = '', order:str = 'asc', limit:int = 1000, offset:int = 0):
        '''
        Get all employees by office.

        Returns a list of tupples \n
        `employee_type` can be `all`, `regular`, or `job order`
        `order_by` can be `id`, `name`, `position`, `office`, `registered`, `time_in`, `time_out` \n
        `order` can be `asc` or `desc`, default to `asc`.
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employees_by_office(office, employee_type = employee_type, order_by = order_by, order = order, limit = limit, offset = offset)

    def employee_exists(self, employee_id:int|str):
        '''
        Check if an employee exist
        
        Returns `true` if exist in database else `false`
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.employee_exists(employee_id)
    
    def get_employee(self, employee_id:int|str):
        '''
        Get employee data

        Returns a tupple containing employee data \n
        (id, name, position, office, registered, time_in, time_out) \n
        Returns `None` if not found
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employee(employee_id)
    
    def get_employee_am_in(self, employee_id:int|str):
        '''
        Get employee regular am in

        Returns a time-formatted string if found otherwise `None`
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employee_am_in(employee_id)
    
    def get_employee_am_out(self, employee_id:int|str):
        '''
        Get employee regular am out

        Returns a time-formatted string if found otherwise `None`
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employee_am_out(employee_id)

    def get_employee_pm_in(self, employee_id:int|str):
        '''
        Get employee regular pm in

        Returns a time-formatted string if found otherwise `None`
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employee_pm_in(employee_id)
    
    def get_employee_pm_out(self, employee_id:int|str):
        '''
        Get employee regular pm out

        Returns a time-formatted string if found otherwise `None`
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employee_pm_out(employee_id)
    
    def get_employee_positions(self, order:str = 'asc'):
        '''
        Get all distinct positions

        Returns a list of all distinct positions \n
        `order` can be `asc` or `desc`, default to `asc`. If order is not valid returns an emply list 
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employee_positions(order = order)
    
    def get_offices(self, order:str = 'asc'):
        '''
        Get all distinct offices

        Returns a list of all distinct offices \n
        `order` can be `asc` or `desc`, default to `asc`. If order is not valid returns an emply list 
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_offices(order = order)
    
    def search_employee(self, by:str, value:str, condition:str = '', employee_type:str = 'all'):
        '''
        Search for employees (uses LIKE operator)

        Returns a list of searched employees \n

        `by` - the column name \n
        `value` - value to be searched \n
        `condition` - combination of special characters\n
        A placeholder for value (value) should be set. \n
        `(e.g.) condition = '%value%'`
        `type` - can be `all`, `regular`, `job order`. Default to all.
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.search_employee(by, value, condition = condition, employee_type = employee_type)
    
    def get_employee_dtrs(self, employee_id:int|str, start:str = '', end:str = '', order_by:str = 'date', order:str = 'desc'):
        '''
        Get DTRs related to `employee_id`

        Returns a list of DTRs found \n
        If `start` and `end` is omitted, returns all DTRs \n
        If `start` is only defined, returns DTRs from that point forward \n
        If `end` is only defined, returns DTRs until that point forward \n
        `order_by` can be `date` or `locked`. Default to date \n
        `order` can be `asc` or `desc`, default to `asc`. If order is not valid returns an emply list 
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employee_dtrs(employee_id, start = start, end = end, order_by = order_by, order = order)
    
    def get_employee_dtr(self, employee_id:int|str, date:str|datetime.date):
        '''
        Get DTR of `employee_id` on specific date
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employee_dtr(employee_id, date)
    
    def check_dtr_date_exists(self, employee_id:int|str, date:str):
        '''
        Check if an employee has existing dtr on the `date` provided

        Returns `True` if dtr exist, otherwise `False`.
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.check_dtr_date_exists(employee_id, date)
    
    def get_employee_dtr_by_date(self, employee_id:int|str, date:str):
        '''
        Check if an employee has existing dtr on the `date` provided

        Returns `True` if dtr exist, otherwise `False`.
        '''
        return self.__database.get_employee_dtr_by_date(employee_id, date)
    
    def get_dtr(self, id:int|str):
        '''
        Get DTR by id
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_dtr(id)
    
    def get_dtr_years(self):
        '''
        Get years with recorded DTR

        Returns a `list` of years
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_dtr_years()
    
    def get_employee_dtr_years(self, employee_id:int|str):
        '''
        Get years with recorded DTR of specific employee

        Returns a `list` of years
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employee_dtr_years(employee_id)

    def get_employees_count(self):
        '''
        Get total employees count
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employees_count()
    
    def get_employees_today_count(self):
        '''
        Get count of employees who clocked in today
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employees_today_count()
    
    def get_registered_employees_count(self):
        '''
        Get count of registered employees
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_registered_employees_count()
    
    def get_employees_today(self):
        '''
        Get employees  who clocked in today
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employees_today()

    def get_employees_from_date(self, date:str|datetime.date):
        '''
        Get employees who clocked in today
        Returns an exception if date are not valid
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_employees_from_date(date = date)

    def get_latest_dtr_date(self):
        '''
        Get latest date of imported DTR
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_latest_dtr_date()

    def get_lates(self, id:int|str, start:datetime.date|str, end:datetime.date|str):
        '''
        Get number of minutes late of employee between two date
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
    
        return self.__database.get_lates(id, start, end)
    
    def get_absents(self, id:int|str, start:datetime.date|str, end:datetime.date|str):
        '''
        Get number of absent days of employee between two dates
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_absents(id, start, end)

    def get_whole_days(self, id:int|str, start:datetime.date|str, end:datetime.date|str):
        '''
        Get number of whole days of employee worked between two dates
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_whole_days(id, start, end)
    
    def get_half_days(self, id:int|str, start:datetime.date|str, end:datetime.date|str):
        '''
        Get number of half days of employee worked between two dates
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_half_days(id, start, end)
    
    def get_job_order_employee_count(self):
        '''
        Get number of job order employees
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_job_order_employee_count()
    
    def get_regular_employee_count(self):
        '''
        Get number of job order employees
        '''
        if not self.__logged:
            raise Exception('Insufficient access privileges')
        
        return self.__database.get_regular_employee_count()
